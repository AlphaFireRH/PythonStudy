#!use/bin/env python3
# -*- coding utf-8 -*-

import openpyxl


# input info
def PleaseInputPath():
    info = 'please input target path: \n'
    return info


# get input
def GetInput():
    return input(PleaseInputPath())


# 创建xlsx文件
def CreateNewXlsx(targetPath):
    wb = openpyxl.Workbook()
    wb.save(targetPath + ".xlsx")
    return wb


# 创建xlsx文件
def CreateNewXlsxFullName(targetPath):
    wb = openpyxl.Workbook()
    wb.save(targetPath)
    return wb


# 打开xlsx文件
def OpenXlsx(targetPath):
    wb = openpyxl.load_workbook(targetPath + ".xlsx")
    return wb


# 打开xlsx文件
def OpenXlsxFullName(targetPath):
    wb = openpyxl.load_workbook(targetPath)
    return wb


# 写xlsx文件
def WriteXlsx(wb, targetPath):
    wb.save(targetPath + ".xlsx")


# 写xlsx文件
def WriteXlsxFullName(wb, targetPath):
    wb.save(targetPath)


# 获取所有Sheet名字
def GetAllSheets(wb):
    return wb.sheetnames


# 获取当前激活的页面
def GetNowActiveSheet(wb):
    return wb.active


# 获取指定名字的页面
def GetTargetSheet(wb, targetSheetName):
    return wb.get_sheet_by_name(targetSheetName)


# 获取指定名字的页面（名称有中文要使用unicode）
def GetTargetSheetCH(wb, targetSheetName):
    return wb.get_sheet_by_name(targetSheetName.encode('unicode'))


# 获取指定名字的页面（名称有中文要使用unicode）
def GetTargetSheet1(wb, targetSheetName):
    return wb[targetSheetName]


# 获取指定名字的页面（名称有中文要使用unicode）
def GetTargetSheetCH1(wb, targetSheetName):
    return wb[targetSheetName.encode('unicode')]


# 获取指定索引的页面
def GetIndexSheet(wb, index):
    return wb.wordsheets[index]


# 创建Sheet
def CreateSheet(wb, targetName):
    wb.create_sheet(targetName)


# 创建中文Sheet
def CreateSheetCH(wb, targetName):
    wb.create_sheet(targetName.encode('unicode'))


# 重命名Sheet
def RenameSheet(ws, targetName):
    ws.title = targetName


# 重命名 中文Sheet
def RenameSheetCH(ws, targetName):
    ws.title = targetName.encode('unicode')


# 根据名字 删除Sheet
def DeleteSheet(wb, targetName):
    wb.remove_sheet(GetTargetSheet(wb, targetName))


# 根据中文名字 删除Sheet
def DeleteSheetCH(wb, targetName):
    wb.remove_sheet(GetTargetSheetCH(wb, targetName))


# 根据索引 删除Sheet
def DeleteSheetIndex(wb, index):
    wb.remove_sheet(GetIndexSheet(wb, index))


# 获取 指定单元内容 ws['A1']
def GetTargetCell(ws, pos):
    return ws[pos]
	
# 获取 指定单元内容 ws['A1:AA']
def GetTargetCells(ws, info):
    return ws[info]

# 获取 指定单元内容 ws['A1:AA']
def GetTargetCellPos(ws, i,j):
    return ws.cell(row=i, column=j)

	

# 设置 指定单元内容 ws['A1']
def SetTargetCellPos(ws, pos, targetValue):
    ws[pos].value = targetValue
    return ws


# 设置 指定单元内容 ws['A1']
def SetTargetCellXY(ws, targetRow, targetColumn, targetValue):
    ws.cell(row=targetRow, column=targetColumn).value = targetValue
    return ws


# 获取最大行
def GetMaxRow(ws):
    return ws.max_row


# 获取最大列
def GetMaxColumn(ws):
    return ws.max_column
